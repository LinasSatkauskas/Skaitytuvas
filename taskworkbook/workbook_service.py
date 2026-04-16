import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

from .constants import FIXED_SHEETS, SHEET_FILES, SHEET_HEADERS


class WorkbookService:
    def __init__(self, excel_path, weekday_parser, stringify):
        self.excel_path = excel_path
        self.workbook = None
        self.weekday_parser = weekday_parser
        self.stringify = stringify

    def load_or_create(self):
        if os.path.exists(self.excel_path):
            self.workbook = load_workbook(self.excel_path)
        else:
            self.workbook = Workbook()
            self.workbook.active.title = FIXED_SHEETS[0]
            for name in FIXED_SHEETS[1:]:
                self.workbook.create_sheet(name)

        self.ensure_fixed_sheets()
        self.order_fixed_sheets()
        self.ensure_headers()
        self.workbook.active = 0
        self.workbook.save(self.excel_path)
        self.export_fixed_sheet_files()
        return self.workbook

    def ensure_fixed_sheets(self):
        existing = set(self.workbook.sheetnames)
        for name in FIXED_SHEETS:
            if name not in existing:
                self.workbook.create_sheet(name)

    def order_fixed_sheets(self):
        ordered = [self.workbook[name] for name in FIXED_SHEETS if name in self.workbook.sheetnames]
        extras = [sheet for sheet in self.workbook.worksheets if sheet.title not in FIXED_SHEETS]
        self.workbook._sheets = ordered + extras

    def ensure_headers(self):
        for sheet_name in FIXED_SHEETS:
            sheet = self.workbook[sheet_name]
            headers = SHEET_HEADERS[sheet_name]
            current_headers = [sheet.cell(row=1, column=index + 1).value for index in range(sheet.max_column)]

            if sheet_name == "Tasks" and "DayOfWeek" not in current_headers:
                sheet.insert_cols(4, 1)
                current_headers = [sheet.cell(row=1, column=index + 1).value for index in range(sheet.max_column)]

            if sheet_name == "CompletedTasks" and "DayOfWeek" not in current_headers:
                sheet.insert_cols(5, 1)
                current_headers = [sheet.cell(row=1, column=index + 1).value for index in range(sheet.max_column)]

            if sheet_name == "TaskHistory" and "DayOfWeek" not in current_headers:
                sheet.insert_cols(5, 1)
                current_headers = [sheet.cell(row=1, column=index + 1).value for index in range(sheet.max_column)]

            if sheet_name == "CompletedTasks" and tuple(current_headers[: len(headers)]) != headers:
                self._reorder_sheet_columns(sheet, headers)

            if sheet_name == "TaskHistory" and tuple(current_headers[: len(headers)]) != headers:
                self._reorder_sheet_columns(sheet, headers)

            for index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=index, value=header)

            if sheet_name == "Tasks":
                for row_number in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row_number, column=4).value is None:
                        sheet.cell(row=row_number, column=4, value=None)

            if sheet_name == "CompletedTasks":
                for row_number in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row_number, column=5).value is None:
                        sheet.cell(row=row_number, column=5, value=None)

            if sheet_name == "TaskHistory":
                for row_number in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row_number, column=5).value is None:
                        sheet.cell(row=row_number, column=5, value=None)

    def _reorder_sheet_columns(self, sheet, desired_headers):
        current_headers = [sheet.cell(row=1, column=index + 1).value for index in range(sheet.max_column)]
        source_lookup = {header: index + 1 for index, header in enumerate(current_headers) if header}
        max_row = sheet.max_row
        max_col = max(sheet.max_column, len(desired_headers))

        for row_number in range(1, max_row + 1):
            row_values = {}
            for header in desired_headers:
                source_column = source_lookup.get(header)
                row_values[header] = sheet.cell(row=row_number, column=source_column).value if source_column else None

            for column_number, header in enumerate(desired_headers, start=1):
                sheet.cell(row=row_number, column=column_number, value=row_values[header])

            for column_number in range(len(desired_headers) + 1, max_col + 1):
                sheet.cell(row=row_number, column=column_number, value=None)

    def export_fixed_sheet_files(self):
        folder = os.path.dirname(self.excel_path)
        for sheet_name in FIXED_SHEETS:
            export_path = os.path.join(folder, SHEET_FILES[sheet_name])
            export_workbook = Workbook()
            export_sheet = export_workbook.active
            export_sheet.title = sheet_name

            source_sheet = self.workbook[sheet_name]
            for row in source_sheet.iter_rows():
                for cell in row:
                    export_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

            export_workbook.save(export_path)

    def save(self):
        self.workbook.save(self.excel_path)
        self.export_fixed_sheet_files()

    def weekday_from_value(self, value):
        if isinstance(value, datetime):
            return value.strftime("%A")

        text = self.stringify(value)
        if not text:
            return ""

        for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(text, pattern).strftime("%A")
            except ValueError:
                continue
        return ""
